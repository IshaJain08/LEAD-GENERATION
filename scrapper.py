from tkinter import *
from instabot import Bot
from selenium.webdriver.common.keys import Keys
from time import sleep
import datetime
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import selenium.common.exceptions as exceptions
import time
import pandas as pd
import pyautogui
import smtplib
from tkinter import filedialog
import openpyxl as openpyxl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def open_linkedin():
    def scrapping():
        linkedin_username = username_login_entry.get()
        linkedin_password = password__login_entry.get()

        def linkedin_login():
            global driver
            options = webdriver.ChromeOptions()
            options.add_argument("start-maximized")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("detach", True)
            # options.add_experimental_option('useAutomationExtension', False)
            try:
                # driver = webdriver.Chrome(chrome_options=options, executable_path='chromedriver.exe')
                driver = webdriver.Chrome(r'C:\Users\isha.jain\Downloads\chromedriver_win32\chromedriver.exe')
                driver.get("https://www.linkedin.com/login")
                WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "username"))).send_keys(
                    linkedin_username)
                driver.find_element_by_id("password").send_keys(linkedin_password)
                driver.find_element_by_xpath("//button[contains(text(), 'Sign in')]").click()
                time.sleep(3)
            except ImportError:
                print("Closing program.")

        search_term = search.get()
        search_term = search_term.split()
        linkedin_search = search_term[0] + "%20" + search_term[1]
        linkedin_search_baseurl = "https://www.linkedin.com/search/results/people/?keywords="
        linkedin_login()
        driver.get(linkedin_search_baseurl + linkedin_search)
        # listOfSearchResult=driver.find_elements_by_xpath('//span[@class ="entity-result__title-text  t-16"]')
        time.sleep(4)

        # load data into a DataFrame object:
        df = pd.DataFrame(columns=["name", "email", "location"])

        numOfEntries = 4  # means 40 entries

        for z in range(numOfEntries):
            for i in range(10):
                time.sleep(5)
                crap2 = driver.find_elements_by_xpath(
                    '//ul[@class ="reusable-search__entity-results-list list-style-none"]')
                li1 = crap2[0].find_elements_by_tag_name("li")
                temp = li1[i].text.split('\n')
                name = temp[0]
                location = ''
                if len(temp) > 1:
                    location = temp[-2] + ' ' + temp[-1]

                print(temp)
                li1[i].click()
                time.sleep(5)
                contactInfoButton = driver.find_elements_by_xpath('//a[@data-control-name ="contact_see_more"]')
                contactInfoButton[0].click()
                time.sleep(3)
                naamSochnaHe = driver.find_elements_by_xpath(
                    '//div[@class ="pv-profile-section__section-info section-info"]')
                crap1 = naamSochnaHe[0].find_elements_by_tag_name("a")

                email = ''

                for j in crap1:
                    print(j.text)
                    if '@' in j.text:
                        email = j.text
                    else:
                        email = 'none'
                df = df.append({'name': name, 'email': email, 'location': location}, ignore_index=True)
                print(df)
                df.to_csv('out.csv', index=False)
                driver.back()
                driver.back()
                time.sleep(5)

            if z < numOfEntries - 1:  # code for clicking next button
                time.sleep(20)
                nextButton = driver.find_elements_by_xpath('//button[@class="artdeco-pagination__button artdeco-pagination__button--next artdeco-button artdeco-button--muted artdeco-button--icon-right artdeco-button--1 artdeco-button--tertiary ember-view"]')
                nextButton[0].click()
                time.sleep(5)


    def photoUpload1():
        filename = filedialog.askopenfilename(initialdir="/", title="Select a File",
                                              filetypes=(("jpeg", "*.jpg"), ("png", "*.png"), ("all files", "*.*")))
        label_photo_path.config(text=filename)

    def post():
        # specifies the path to the chromedriver.exe
        driver = webdriver.Chrome(r'C:\Users\isha.jain\Downloads\chromedriver_win32\chromedriver.exe')

        # reading csv file

        # reading username
        myUsername = username_login_entry.get()
        myPassword = password__login_entry.get()

        driver.get("https://www.linkedin.com/login")
        WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "username"))).send_keys(
            myUsername)
        driver.find_element_by_id("password").send_keys(myPassword)
        driver.find_element_by_xpath("//button[contains(text(), 'Sign in')]").click()
        time.sleep(3)

        # reading post
        myPost = caption_entry.get(1.0, END)
        path_to_image = label_photo_path['text']
        # label_photo_path['text']
        imageButton = driver.find_elements_by_xpath('//button[@data-control-name ="share.select_image"]')

        # locate post field by_xpath
        postButton = driver.find_elements_by_xpath(
            '//button[@class ="artdeco-button artdeco-button--muted artdeco-button--4 artdeco-button--tertiary share-box-feed-entry__trigger--v2"]')

        # .click() to mimic button click
        postButton[0].click()

        # locate text editor by_class_name
        postContent = driver.find_element_by_class_name('ql-editor')

        # send_keys() to simulate key strokes
        postContent.send_keys(myPost)
        time.sleep(5)
        imageButton = driver.find_elements_by_xpath('//button[@data-control-name ="share.select_image"]')
        # .click() to mimic button click
        imageButton[0].click()

        uploadbox = driver.find_elements_by_xpath('//input[@id="image-sharing-detour-container__file-input"]')
        uploadbox[0].send_keys(path_to_image)

        submitImageButton = driver.find_elements_by_xpath('//button[@data-control-name ="confirm_selected_photo"]')
        # .click() to mimic button click
        submitImageButton[0].click()

        # locate post button by_id
        post = driver.find_elements_by_xpath(
            '//button[@class ="share-actions__primary-action artdeco-button artdeco-button--2 artdeco-button--primary ember-view"]')

        # .click() to mimic button click
        post[0].click()

        pyautogui.hotkey('alt', 'f4')

    root1 = Tk()
    root1.geometry('700x520')
    root1.resizable(width=0, height=0)
    root1.title("Linked In")
    text = Label(root1, text="Linked In", font=('Arial', 18, 'bold'))
    text.pack(padx=30, pady=30)

    # Middle Part
    user = Label(root1, text="User Id")
    user.pack()
    user.place(x=250, y=80)
    username_login_entry = Entry(root1, textvariable="username", width=30)
    username_login_entry.pack()
    username_login_entry.place(x=330, y=80)

    pwd = Label(root1, text="Password")
    pwd.pack()
    pwd.place(x=250, y=110)

    password__login_entry = Entry(root1, textvariable="password", show='*')
    password__login_entry.pack()
    password__login_entry.place(x=330, y=110)

    # Left Part

    label_search = Label(root1, text="Search Term")
    label_search.pack()
    label_search.place(x=90, y=180)
    search = Entry(root1, textvariable="search")
    search.pack()
    search.place(x=160, y=180)

    btn1 = Button(root1, text='Download', width=12, command=scrapping)
    btn1.pack()
    btn1.place(x=150, y=280)

    # Right Part Area

    photo = Label(root1, text="Photo", justify=LEFT)
    photo.pack()
    photo.place(x=400, y=180)

    button_to = Button(root1, text="Upload Photo", command=photoUpload1, width=16)
    button_to.pack()
    button_to.place(x=460, y=178)

    label_photo_path = Label(root1,text="", fg="blue")
    label_photo_path.place(x=460, y=210)

    caption = Label(root1, text="Caption")
    caption.pack()
    caption.place(x=400, y=280)

    caption_entry = Text(root1, height=4, width=20)
    caption_entry.pack()
    caption_entry.place(x=460, y=280)

    Label(root1, text="").pack()
    btn = Button(root1, text='Post', width=12, command=post)
    btn.pack()
    btn.place(x=490, y=380)

    root1.mainloop()

def open_insta():
    root2 = Tk()

    def photoUpload():
        filename = filedialog.askopenfilename(initialdir="/", title="Select a File",
                                              filetypes=(("jpeg", "*.jpg"), ("png", "*.png"), ("all files", "*.*")))
        label_photo_path.config(text=filename)

    def post():

        username = username_login_entry.get()
        password = password__login_entry.get()  # your IG password
        photo_path = label_photo_path['text']
        caption1 = caption.get()

        driver = webdriver.Chrome(r'C:\Users\isha.jain\Downloads\chromedriver_win32\chromedriver.exe')
        driver.implicitly_wait(5)
        driver.get('https://www.instagram.com/')
        # -------login process starts
        # finding input boxes for username and password and pasing the appropriate values
        driver.find_element_by_xpath("//input[@name='username']").send_keys(username)
        driver.find_element_by_xpath("//input[@name='password']").send_keys(password)
        # findind login button and clicking it
        driver.find_element_by_xpath("//button[@type='submit']").click()

        time.sleep(3)

        # 2nd pop-up
        driver.find_element_by_xpath(
            '/html/body/div[4]/div/div/div/div[3]/button[2]').click()
        driver.get('https://www.instagram.com/' + username)
        time.sleep(4)
        bot = Bot()

        bot.login(username=username,
                  password=password)

        bot.upload_photo(photo_path, caption=caption1)

    def message():
        username = username_login_entry.get()
        password = password__login_entry.get()  # your IG password
        message1 = message_entry.get('1.0', END)

        count = 100  # number of profiles you want to get
        account = username  # account from username ur
        page = "follower"  # from following or followers

        driver = webdriver.Chrome(r'C:\Users\isha.jain\Downloads\chromedriver_win32\chromedriver.exe')
        driver.implicitly_wait(5)
        driver.get('https://www.instagram.com/')
        # -------login process starts
        # finding input boxes for username and password and pasing the appropriate values
        driver.find_element_by_xpath("//input[@name='username']").send_keys(username)
        driver.find_element_by_xpath("//input[@name='password']").send_keys(password)
        # findind login button and clicking it
        driver.find_element_by_xpath("//button[@type='submit']").click()


        # 2nd pop-up
        driver.find_element_by_xpath(
            '/html/body/div[4]/div/div/div/div[3]/button[2]').click()
        time.sleep(4)

        driver.get('https://www.instagram.com/%s' % account)
        sleep(2)
        driver.find_element_by_xpath('//a[contains(@href, "%s")]' % page).click()
        scr2 = driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/header/section/ul/li[2]/a')
        sleep(2)
        text1 = scr2.text
        t = text1.split()
        t2 = int(t[0]) + 1
        print(text1)
        x = datetime.datetime.now()
        print(x)
        li = []
        for i in range(1, t2):
            scr1 = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]/ul/div/li[%s]' % i)
            driver.execute_script("arguments[0].scrollIntoView();", scr1)
            sleep(1)
            text = scr1.text
            list = text.split()
            li.append(list[0])
            print(list[0])
            # print('{};{}'.format(i, list[0]))
            # print(i + ";" + list[0])
            if i == (t2 - 1):
                print(x)
        print(li)
        time.sleep(5)
        button = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[1]/div/div[2]/button')
        driver.execute_script("arguments[0].click();", button)
        time.sleep(5)
        # direct button
        driver.find_element_by_xpath(
            '//a[@class="xWeGp"]').click()
        time.sleep(3)

        # clicks on pencil icon
        driver.find_element_by_xpath(
            '//*[@id="react-root"]/section/div/div[2]/div/div/div[1]/div[1]/div/div[3]/button').click()
        time.sleep(2)
        for k in li:
            # enter the username
            driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]/div[1]/div/div[2]/input').send_keys(k)
            time.sleep(2)

            # click on the username
            driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]/div[2]/div[1]/div/div[3]/button').click()
            time.sleep(2)

            # next button
            driver.find_element_by_xpath(
                '/html/body/div[4]/div/div/div[1]/div/div[2]/div/button').click()
            time.sleep(2)

            # click on message area
            send = driver.find_element_by_xpath(
                '//*[@id="react-root"]/section/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/div[2]/textarea')

            # types message
            send.send_keys(message1)
            time.sleep(1)

            # send message
            send.send_keys(Keys.RETURN)
            time.sleep(2)

            # clicks on direct option or pencl icon
            driver.find_element_by_xpath(
                '/html/body/div[1]/section/div/div[2]/div/div/div[1]/div[1]/div/div[3]/button').click()
            time.sleep(2)
        driver.quit()

    root2.geometry('700x520')
    root2.resizable(width=0, height=0)
    root2.title("MyInsta")

    text = Label(root2, text="INSTAGRAM", font=('Arial', 18, 'bold'))
    text.pack(padx=30, pady=30)
    # Middle Part
    user = Label(root2, text="UserId")
    user.pack()
    user.place(x=260, y=80)
    username_login_entry = Entry(root2, textvariable="username")
    username_login_entry.pack()
    username_login_entry.place(x=340, y=80)

    # pwd=Label(root2, text="").pack()

    pwd = Label(root2, text="Password")
    pwd.pack()
    pwd.place(x=260, y=110)

    password__login_entry = Entry(root2, textvariable="password", show='*')
    password__login_entry.pack()
    password__login_entry.place(x=340, y=110)

    # Left Part

    photo = Label(root2, text="Photo", justify=LEFT)
    photo.pack()
    photo.place(x=100, y=180)

    button_to = Button(root2, text="Upload Photo", command=photoUpload, width=16)
    button_to.pack()
    button_to.place(x=160, y=178)

    label_photo_path = Label(root2, text="", fg="blue")
    label_photo_path.place(x=160, y=210)

    cptn = Label(root2, text="Caption")
    cptn.pack()
    cptn.place(x=100, y=280)
    caption = Entry(root2, textvariable="caption")
    caption.pack()
    caption.place(x=160, y=280)

    btn1 = Button(root2, text='Post', width=12, command=post)
    btn1.pack()
    btn1.place(x=150, y=380)

    # Right Part Area

    mssg = Label(root2, text="Message")
    mssg.pack()
    mssg.place(x=400, y=180)

    message_entry = Text(root2, height=10, width=26)
    message_entry.pack()
    message_entry.place(x=460, y=180)

    Label(root2, text="").pack()
    btn = Button(root2, text='Send', width=12, command=message)
    btn.pack()
    btn.place(x=490, y=380)

    root2.mainloop()

def open_email():
    def Take_input():
        lstatus.config(text="Sending.. !!")
        sname1 = efrom.get()
        sub = esub.get()
        body = ebody.get(1.0,END)
        to_path = l_file_to['text']
        pswd = epswd.get()
        attach_path = l_file_attach['text']
        wb = openpyxl.load_workbook(to_path)
        sheet = wb.active
        mc = sheet.max_column
        mr = sheet.max_row
        col = 0
        for i in range(1, 3):

            name = sheet.cell(row=1, column=i).value
            if name == "email":
                col = i
                break

        for j in range(2, mr + 1):
            send_mail(sname1, pswd, sheet.cell(row=j, column=col).value, attach_path, sub, body)
        lstatus.config(text="Mail Sent !!")

    def browseFiles_attach():
        filename = filedialog.askopenfilename(initialdir="/", title="Select a File", )

        l_file_attach.config(text=filename)
        to_path = filename

    def browseFiles_to():
        filename = filedialog.askopenfilename(initialdir="/", title="Select a File",
                                              filetypes=(("Excel", "*.xlsx"), ("all files", "*.*")))
        l_file_to.config(text=filename)
        to_path = filename

    window = Tk()
    window.geometry("700x520")
    window.title("Email Automation")
    text = Label(window, text="Email Automation", font=('Arial', 18, 'bold'))
    text.pack(padx=30, pady=30)

    lfrom = Label(window, text="From").place(x=220, y=120)
    lpswd = Label(window, text="Password").place(x=220, y=160)
    lto = Label(window, text="To").place(x=220, y=200)
    lsub = Label(window, text="Subject").place(x=220, y=240)
    lbody = Label(window, text="Body").place(x=220, y=280)
    lattach = Label(window, text="Attachment").place(x=220, y=350)
    lstatus = Label(window, text="", font="10",fg="red")

    efrom = Entry(window,width=25)
    epswd = Entry(window, show='*')
    # to
    button_to = Button(window, text="Sender's Emails File", command=browseFiles_to)
    l_file_to = Label(window, text="", width=30, fg="blue",anchor="w")

    esub = Entry(window)
    ebody = Text(window,height=3,width=20)
    # attach
    l_file_attach = Label(window, text="", width=30, fg="blue",anchor="w")
    button_attach = Button(window, text="Browse Files", command=browseFiles_attach)

    send = Button(window, height=2, width=20, text="Send Mail", command=lambda: Take_input())

    efrom.place(x=350, y=120)
    epswd.place(x=350, y=160)

    button_to.place(x=350, y=200)
    l_file_to.place(x=480, y=200)

    esub.place(x=350, y=240)
    ebody.place(x=350, y=280)

    button_attach.place(x=350, y=350)
    l_file_attach.place(x=430, y=350)

    send.place(x=250, y=390)
    lstatus.place(x=300, y=440)

    def send_mail(sname, pswd, rname, filename, subject, body):
        msg = MIMEMultipart()
        msg['From'] = sname
        msg['To'] = rname
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        file = filename
        boo=file.split("/")
        file1=file.replace(filename,boo[-1])
        attachment = open(file, "rb")
        p = MIMEBase('application', 'octet-stream')
        p.set_payload(attachment.read())
        encoders.encode_base64(p)
        p.add_header('Content-Disposition', "attachment; filename= " + file1)
        msg.attach(p)

        try:
            server = smtplib.SMTP("smtp.gmail.com", 587)

            server.starttls()
            server.login(sname, pswd)
            text = msg.as_string()
            server.sendmail(sname, rname, text)
            server.quit()
        # TODO: Send email here
        except Exception as e:
            # Print any error messages to stdout
            print(e)

    window.mainloop()

root = Tk()
root.geometry('700x520')
root.resizable(width=0, height=0)
root.title("Title(to be changed later)")
text1=Label(root,text="Tech Expo 2021",font=('Arial',25,'bold'))
text1.pack(padx=30,pady=30)

text=Label(root,text="Lead Generation (Automation)",font=('Arial',18,'bold'))
text.pack(padx=30,pady=30)

btn1 = Button(root,text = "Linked In",command=open_linkedin,font =('calibri', 15, 'bold'),width=15,height=2)
btn1.place(x=60,y=261)

btn2 = Button(root,text = "Instagram",command=open_insta,font =('calibri', 15, 'bold'),width=15,height=2)
btn2.place(x=280,y=260)

btn3 = Button(root,text = "Email",command=open_email,font =('calibri', 15, 'bold'),width=15,height=2)
btn3.place(x=500,y=260)

root.mainloop()


